
import os

try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl import Workbook

# 文言文件目录
rootdir = 'D:/HYRON_input_AI_190308/FZ-Application/xml'

# output 位置
resultDir = 'Result'
xlsxFile = 'xml2xlsx.xlsx'
xmlFile = '_UnitData.xml'
NULLAttrib = '%NULL%'

# 特殊字符
quot ='&quot;'

# 外部参照的属性一览
procItemDialog = 'ProcItemDialog.'
settingType = 'settingType'

unitAttrList = ['no','ident','isLoopStart','isLoopEnd','setEnable',\
        'getEnable','multiData','multiNum','multiCalc',\
        'analyzable','LowerDataIdent','UpperDataIdent','inputParam',\
        'outputParam','isGroup','isSubGroup','settingType','isAction']
# 外部参照的设定对话框
diagList = ['CalculationDialog','ColorDialog','ColorExtractDialog','FigureDialog',\
        'PointDialog','UnitSelectDialog','CommonDialog','ValueSetDialog','LimitSetDialog',\
        'ComboBoxDialog','NoDialog','FileDialog','FolderDialog','MultiLineTextDialog']
# 外部参照的设定对话框的设定属性一览
diagAttrList = [\
    ['MaxLength'],\
    ['ThroughEnable','DisplaySubNo','DataIdentR','DataIdentG','DataIdentB',\
        'ColorRangeSupport','DataIdentGR','DataIdentGG','DataIdentGB','DataIdentColor',\
        'DataIdentColorG','ColorAreaFigNo'],\
    ['ThroughEnable','DisplaySubNo','AutoSetting','DataIdentUpperH','DataIdentLowerH',\
        'DataIdentUpperS','DataIdentLowerS','DataIdentUpperV','DataIdentLowerV','DataIdent'],\
    ['ThroughEnable','DisplaySubNo','FigureNo','FigureKind','FigureMax','CircleMode',\
        'LineArrowMode','NotOnlyFigureEnabled','OrNotEnabled','DeleteDisable',\
        'ActionType','ActionIdent','ActionValue'],\
    ['ThroughEnable','DisplaySubNo','DecimalPlaces','MaxX','MinX','MaxY','MinY','DataIdentX','DataIdentY'],\
    ['TargetScene','MaxUnitNo','MinUnitNo','ItemIdentFilters','CategoryFilters'],\
    ['MaxLength'],\
    ['DecimalPlaces','Max','Min','Interval'],\
    ['SliderEnable','DecimalPlaces','Max','Min','LargeChange','SmallChange',\
        'TickFrequency','Interval','DataIdentUppValue','DataIdentLowValue'],\
    ['MessageFileIdent','TextIdentList','ValueList'],\
    ['ActionType','ActionIdent','ActionValue'],\
    ['RootFolderIdent','MessageFileIdent','FilterMessage','ExSupportFolderList','ModeOpen','IsImage',\
        'IsLocalPC','IsMem','IsUSBMemory','PreviewEnabled','ShowLoggingImage','ShowRegisteredImage','ShowStandardImage'],\
    ['RootFolderIdent','ExSupportFolderList','IsImageFolder','IsLocalPC','IsMemFolder','IsUSBMemory','VisibleNewFolder'],\
    ['MaxLength']\
    ]

# 外部参照的设定对话框属性个数列表
diagAttrbLenList = list(map(lambda x:len(x),diagAttrList))

def getDiagOffset(x):
    '''
        计算指定的设定对话框属性的偏移位置
    '''
    offset = 0
    for idx in range(0,diagList.index(x)): 
        offset += diagAttrbLenList[idx]
    return offset

# 外部参照的设定对话框属性的偏移位置列表
diagAttrbOffList = list(map(getDiagOffset,diagList))

# 外部参照的设定对话框的词典
diagAttrDict = dict(zip(diagList,diagAttrList))

def make_header():
    '''
        结果文件(csv/xlsx)的Header行
    '''
    head1=[None]*(len(unitAttrList)+1)
    for key in diagAttrDict:
        head1.append(key)
        head1 += [None]*(len(diagAttrDict[key]) - 1)

    head2 = ['tag'] + unitAttrList
    for key in diagAttrDict:
        head2 += diagAttrDict[key]
    
    return [head1,head2]

def parseXMLbyElementTree(xmlfile):
    '''
        用ElementTree解析XML文件
    '''
    # 解析XML文件
    tree = ET.parse(xmlfile)  # <class 'xml.etree.ElementTree.ElementTree'>
    root = tree.getroot()     # 获取根节点 <Element 'data' at 0x02BF6A80>
    # print(root.tag, ":", root.attrib) # 打印根元素的tag和属性
    unitdata = root[0]        # 获取unitdata节点
    # print(unitdata.tag, ":", unitdata.attrib) # 打印根元素的tag和属性
    
    err = False
    lines = make_header()
    for node in unitdata:
        # print('node={}'.format(node.tag))

        line = [node.tag]
        for attr in unitAttrList:
            line.append(node.attrib.get(attr,NULLAttrib))
            # print('--','{}={}'.format(attr,node.attrib.get(attr,NULLAttrib)))

        for subnode in node:
            
            if subnode.tag == 'setting':
                diagType = subnode.attrib['Type'].split('.')[1]
                # print('----','Type={}'.format(procItemDialog + diagType))
                if diagType in diagList:
                    line[unitAttrList.index(settingType)+1] = diagType
                    line += [None]*diagAttrbOffList[diagList.index(diagType)]
                    # print('----','AttribPos={}'.format(offset))

                    for diagAttrs in diagAttrDict[diagType]:
                        line.append(subnode.attrib.get(diagAttrs,NULLAttrib))
                        # print('----','{}={}'.format(diagAttrs,subnode.attrib.get(diagAttrs,NULLAttrib)))
                else:
                    print('Type Error %s' % diagType)
                    err = True
                    break
        lines.append(line)

        if err: 
            lines = []
            break

    # print(lines)
    unit = os.path.basename(xmlfile).split('_')[0]
    return lines,unit
 
def make_csv_file(lines, unit_name):
    '''
        行list保存到CSV文件
    '''
    import csv

    dir = os.path.join(os.getcwd(), resultDir)
    if not os.path.isdir(dir):
        os.mkdir(dir)
    fn = os.path.join(dir,'{}.csv'.format(unit_name))

    # CSV
    with open(fn, 'w', encoding='UTF-8-sig', newline='') as fw:
        writer = csv.writer(fw)
        for line in lines:
            writer.writerow(line)

def make_xlsx_file(wb,lines, unit_name):
    '''
        生成xlsx sheet
    '''
    if len(unit_name) > 31: unit_name = unit_name[0:31]
    ws = wb.create_sheet(unit_name)

    for line in lines:
        ws.append(line)

def get_xmlfile_list():
    '''
        得到处理项目xml文件一览
    '''
    import glob # 文件名模式匹配库
    filelist = glob.glob(os.path.join(rootdir,'*%s' % xmlFile))
    return filelist

def process_xmlfiles(filelist, csv=False):
    '''
        处理所有XML文件
    '''
    if not csv:
        # 创建一个工作表
        wb = Workbook()
        # 删除默认sheet
        sheet = wb.active
        wb.remove(sheet)

    err_units = []
    for name in filelist:
        # print('{} start'.format(name))
        lines,unit = parseXMLbyElementTree(name)

        if len(lines) == 0: err_units.append(unit)
        if not csv:
            make_xlsx_file(wb,lines,unit)
        else:
            make_csv_file(lines,unit)        
        
        print('{} finished'.format(os.path.basename(name)))

    if not csv:
        # 保存xlsx文件
        dir = os.path.join(os.getcwd(), resultDir)
        if not os.path.isdir(dir):
            os.mkdir(dir)
        fn = os.path.join(dir,xlsxFile)
        wb.save(fn)
        print('{} saved'.format(os.path.basename(fn)))

    print('all error units: \n', err_units)

def parse_ws2xml_ET(ws):
    '''
        读取Sheet内容，生成ElementTree
    '''
    def None2Str(x):
        '''
            把None转换成''
        '''
        if not x:x=''
        return str(x)
    
    # 初始化xml root
    root = ET.Element('Reference')
    root.text = '\n' + ' ' * 2
    unitdata = ET.SubElement(root,'UNITDATA', {})
    unitdata.text = '\n' + ' ' * 4
    unitdata.tail = '\n'

    for row in list(ws.rows)[2:]:
        # 读取sheet中的一行到列表中
        line = list(map(None2Str,[cell.value for cell in row]))

        # 解析line,返回tag名，tag属性集合，setting属性集合
        tag,tagAttr,setAttr = parse_xls_line(line) 
        # print(tag,tagAttr,setAttr)

        # 生成subElement
        tagNode=ET.SubElement(unitdata,tag,tagAttr)
        tagNode.tail = '\n' + ' ' * 4
        if len(setAttr) != 0:
            tagNode.text = '\n' + ' ' * 6
            setNode = ET.SubElement(tagNode,'setting',setAttr)
            setNode.tail = '\n' + ' ' * 4

    tagNode.tail = '\n' + ' ' * 2  # 修改最后一个tag的缩进
    save_xml_ET(root,ws.title)

def parse_ws2xml_manual(ws):
    '''
        读取Sheet内容，生成文件行
    '''
    def None2Str(x):
        '''
            把None转换成''
        '''
        if not x:x=''
        return str(x)

    xmlLines=['<?xml version="1.0" encoding="UTF-8" standalone="no"?>','<Reference>','  <UNITDATA>']
    xmlLineTails=['  </UNITDATA>','</Reference>']

    for row in list(ws.rows)[2:]:
        # 读取sheet中的一行到列表中
        line = list(map(None2Str,[cell.value for cell in row]))

        # 解析line,返回tag名，tag属性集合，setting属性集合
        tag,tagAttr,setAttr = parse_xls_line(line) 

        attrList=['{}="{}"'.format(k,v) for k,v in tagAttr.items()]

        if len(setAttr) != 0:
            tagLine ='{}<{} {}>'.format(' ' * 4, tag, ' '.join(attrList))
            xmlLines.append(tagLine)

            setAttrList = ['{}="{}"'.format(k,v.replace('"',quot)) for k,v in setAttr.items()]
            setLine = '{}<{} {}/>'.format(' ' * 6, 'setting', ' '.join(setAttrList))
            xmlLines.append(setLine)

            tagLineL ='{}</{}>'.format(' ' * 4, tag)
            xmlLines.append(tagLineL)
        else:
            tagLine ='{}<{} {}/>'.format(' ' * 4, tag, ' '.join(attrList))
            xmlLines.append(tagLine)

    save_xml_manual(xmlLines + xmlLineTails,ws.title)

def parse_xls_line(line):
    '''
        解析行内容，返回tag名，tag属性集合，setting属性集合
    '''
    # attrib名字列表，去除了settingType
    attrNames = unitAttrList[:-2] + [unitAttrList[-1]] 

    # 对应attrib名字列表的attrib值的列表
    attrValues = line[1:len(attrNames)] + [line[len(attrNames)+1]]

    # 节点属性字典,含%NULL%
    tmpattr = dict(zip(attrNames,attrValues))

    # 节点属性字典，去掉%NULL%属性
    tagAttrList = {k: v for k, v in tmpattr.items() if v != NULLAttrib}

    # 节点元素名称
    tagName = line[0]

    # 节点setting Type
    diagName = line[unitAttrList.index(settingType)+1]

    settingAttrDict = {}
    if diagName != NULLAttrib:
        # 取出diag相关的设定值
        allDiagLine = line[len(unitAttrList)+1:]

        diagAttrbNames = diagAttrDict[diagName]
        idx = diagList.index(diagName)
        diagAttrbValues = allDiagLine[diagAttrbOffList[idx]:]
        tmpsetattr = dict(zip(['Type'] + diagAttrbNames, [procItemDialog+diagName] + diagAttrbValues))
        # 节点属性字典，去掉%NULL%属性
        settingAttrDict = {k: v for k, v in tmpsetattr.items() if v != NULLAttrib}

    return tagName,tagAttrList,settingAttrDict

def save_xml_manual(xmlLines,sheetname):
    '''
        保存xml文件
    '''
    # 结果目录
    dir = os.path.join(os.getcwd(), resultDir)
    if not os.path.isdir(dir):
        os.mkdir(dir)
    fn = os.path.join(dir,sheetname + xmlFile)

    # 写xml文件
    with open(fn, 'w', encoding='UTF-8-sig', newline='') as f:
        f.writelines(map(lambda x: x + '\r\n', xmlLines))

    print('{} created'.format(os.path.basename(fn)))

def save_xml_ET(xmlRoot,sheetname):
    '''
        保存xml文件(ElementTree)
    '''
    tree=ET.ElementTree(xmlRoot)

    # 保存xml文件
    dir = os.path.join(os.getcwd(), resultDir)
    if not os.path.isdir(dir):
        os.mkdir(dir)
    fn = os.path.join(dir,sheetname + xmlFile)
    tree.write(fn,encoding='UTF-8',xml_declaration=True)
    print('{} created'.format(os.path.basename(fn)))


def xml2xlsx():
    '''
        xml->xlsx(csv)
    '''
    process_xmlfiles(get_xmlfile_list())

def xlsx2xml():
    '''
        xlsx-xml
    '''    
    fn = os.path.join(os.path.join(os.getcwd(), resultDir),xlsxFile)
    if os.path.isfile(fn):
        # 打开xlsx文件
        wb = load_workbook(fn)
        for sheet in wb.worksheets:
            # 枚举sheet
            # process_ws2xml_ET(sheet)
            parse_ws2xml_manual(sheet)

        wb.close()
        print('All xml files created')

def main():
    """
        主函数
    """
    xml2xlsx()
    xlsx2xml()

if __name__ == '__main__':
    main()
