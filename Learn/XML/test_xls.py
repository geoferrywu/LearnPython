
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def test_xls():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "New Shit"
    sheet['C3'] = 'Hello world!'
    for i in range(10):
        sheet["A%d" % (i+1)].value = i + 2
    # wb.remove(sheet)

    ns = wb.create_sheet('2')
    for i in range(10):
        ns["A%d" % (i+1)].value = i + 2
        ns.append([i]*5)

    ws3 = wb.create_sheet(title="Data")
    for col in range(1, 10):
        for row in range(1, 10):
            ws3.cell(col, row).value = col*row

    wb.save('d:\\保存一个新的excel.xlsx')

def test():
    test_xls()

def main():
    """
        主函数
    """
    test()

if __name__ == '__main__':
    main()